from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.schemas import DailyLeadReport, ScoredLead


class ExcelReportBuilder:
    columns = [
        "Business Name",
        "Industry",
        "Location",
        "Phone Number",
        "Email",
        "Website",
        "Instagram/Facebook",
        "Contact Person",
        "Estimated Organization Size",
        "Operational Complexity Score",
        "Lead Score",
        "SaaS Potential Score",
        "Likely Operational Challenges",
        "Suggested Nexora Entry Point",
        "Suggested Conversation Angle",
        "Notes",
    ]

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def build(self, report: DailyLeadReport) -> str:
        date_slug = report.date.strftime("%Y-%m-%d")
        path = self.output_dir / f"NEXORA_DAILY_LEADS_{date_slug}.xlsx"
        rows = [self._row(lead) for lead in report.leads]
        df = pd.DataFrame(rows, columns=self.columns)
        summary = pd.DataFrame(
            [
                {"Metric": "Total Leads", "Value": len(report.leads)},
                {"Metric": "Schools", "Value": report.school_count},
                {"Metric": "Solar Companies", "Value": report.solar_count},
                {"Metric": "Average Lead Score", "Value": round(df["Lead Score"].mean(), 2) if not df.empty else 0},
                {"Metric": "Report Date", "Value": date_slug},
            ]
        )
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Daily Leads", index=False)
            summary.to_excel(writer, sheet_name="Summary", index=False)
            self._style_leads(writer.book["Daily Leads"])
            self._style_summary(writer.book["Summary"])
        return str(path)

    def _row(self, lead: ScoredLead) -> dict[str, object]:
        return {
            "Business Name": lead.business_name,
            "Industry": "School" if lead.industry.value == "school" else "Solar Company",
            "Location": lead.location,
            "Phone Number": lead.phone_number,
            "Email": lead.email,
            "Website": lead.website,
            "Instagram/Facebook": lead.social_url,
            "Contact Person": lead.contact_person,
            "Estimated Organization Size": lead.estimated_organization_size,
            "Operational Complexity Score": lead.operational_complexity_score,
            "Lead Score": lead.lead_score,
            "SaaS Potential Score": lead.saas_potential_score,
            "Likely Operational Challenges": lead.likely_operational_challenges,
            "Suggested Nexora Entry Point": lead.suggested_nexora_entry_point,
            "Suggested Conversation Angle": lead.suggested_conversation_angle,
            "Notes": lead.notes,
        }

    def _style_leads(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(color="FFFFFF", bold=True)
        alt_fill = PatternFill("solid", fgColor="F3F4F6")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
            for idx in [10, 11, 12]:
                score_cell = row[idx - 1]
                score = int(score_cell.value or 0)
                score_cell.fill = PatternFill("solid", fgColor="DCFCE7" if score >= 75 else "FEF3C7" if score >= 55 else "FEE2E2")
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 55)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    def _style_summary(self, ws) -> None:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="0F766E")
            cell.font = Font(color="FFFFFF", bold=True)
        for column_cells in ws.columns:
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = 26
